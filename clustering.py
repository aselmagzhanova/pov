import openpyxl
import random
import numpy as np
import pandas as pd


#загрузка данных из xlsx файла
def data_download():
    wb = openpyxl.load_workbook(filename='input.xlsx')
    sheet = wb['Лист1']
    rows = sheet.iter_rows()
    data_gen = ([cell.value for cell in row] for row in rows)
    data = []

    for i in data_gen:
        data.append(i[1::])

    data = data[1::]  # убираем название для удобства рассчетов
    return data

#начальное задание центроидов
def cluster_centroid(data):
    n = int(input("Введите количество кластеров:"))
    cl = []
    n_param = len(data[1])
    for i in range(n):
        cl.append([])
        for j in range(n_param):
            cl[i].append(random.uniform(0,1))
    return  cl

#рассчет расстояний до центроидов
def cluster_distance(data, clusters):
    cluster_quantity = len(clusters)
    # промежуточные суммы для вычислений расстояний до центроидов кластеров
    sum = []

    # списки, содержащие расстояний до кластеров
    cl_s = []

    # обнуляем список сумм
    for i in range(len(data)):
        sum.append([])
        for j in range(cluster_quantity):
            sum[i].append(0)

    #обнуляем список расстояний до кластеров
    for i in range(len(data)):
        cl_s.append([])
        for j in range(cluster_quantity):
            cl_s[i].append(0)

    #высчитываем квардраты расстояний
    for i in range(len(data)):
        for j in range(len(data[i])):
            for k in range(cluster_quantity):
                sum[i][k] += (clusters[k][j] - data[i][j])**2


    #высчитываем расстояния
    for i in range(len(data)):
        for k in range(cluster_quantity):
            cl_s[i][k] = sum[i][k]**(0.5)

    return cl_s


#разбиение на кластеры
def clustering(data, distance):
    #список кластеров
    cluster_list = []
    # записываем в список кластеров
    for i in range(len(data)):
        cluster_list.append(distance[i].index(min(distance[i])))
    return cluster_list


#вычисление новых координат центроидов
def calculate_new_clusters(data, clusters, cluster_list):
    sum_sr = []
    n_sr = []

    # обнуляем сумму параметров для рассчета нового центроида
    for i in range(len(clusters)):
        sum_sr.append([])
        for j in range(len(clusters[i])):
            sum_sr[i].append(0)

    # обнуляем кол-во парамеров для рассчета нового кластера
    for i in range(len(clusters)):
        n_sr.append([])
        for j in range(len(clusters[i])):
            n_sr[i].append(0)


    # высчитываем суммы и кол-ва параметров для новых рассчетов
    for i in range(len(data)):
        for j in range(len(data[i])):
            sum_sr[cluster_list[i]][j] += data[i][j]
            n_sr[cluster_list[i]][j] += 1

    for i in range(len(clusters)):
        for j in range(len(clusters[i])):
            if(n_sr[i][j] == 0):
                n_sr[i][j] = 1

    # вычисление новых координат
    for i in range(len(clusters)):
        for j in range(len(clusters[i])):
            sum_sr[i][j] = sum_sr[i][j] / n_sr[i][j]

    return sum_sr

#нормализация первоначальных данных
def normalize(data):
    data = np.array(data)
    data = data.transpose()
    norm_data = []
    for i in range(len(data)):
        norm_data.append([])
        for j in range(len(data[i])):
            norm_data[i].append((data[i][j] - min(data[i]))/(max(data[i]) - min(data[i])))
    norm_data = np.array(norm_data)
    norm_data = norm_data.transpose()

    return norm_data

#вывод данных в xlsx файл
def data_output(clusters):
    wb = openpyxl.load_workbook(filename='input.xlsx')
    sheet = wb['Лист1']
    rows = sheet.iter_rows()
    data_gen = ([cell.value for cell in row] for row in rows)
    data = []  # список из гнератора

    for i in data_gen:
        data.append(i[0])
    data = data[1::]
    for i in range(len(clusters)):
        clusters[i] += 1

    data = np.array([data, clusters])
    data = data.transpose()

    ## convert your array into a dataframe
    df = pd.DataFrame(data)

    ## save to xlsx file

    filepath = 'output.xlsx'

    df.to_excel(filepath, index=False)
    return data


if __name__ == '__main__':

    data = data_download()
    data = normalize(data)
    clusters = cluster_centroid(data)
    distance = cluster_distance(data, clusters)
    cluster_list = clustering(data, distance)
    new = calculate_new_clusters(data, clusters, cluster_list)
    new1 = []
    f = True
    while(f != False):
        distance = cluster_distance(data, new)
        cluster_list = clustering(data, distance)
        new1 = calculate_new_clusters(data, new, cluster_list)

        if (new1 != new):
            new = new1
        else:
            f = False

    data_o = data_output(cluster_list)
    print(data_o)
